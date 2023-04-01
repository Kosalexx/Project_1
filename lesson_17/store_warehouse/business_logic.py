from time import time
from datetime import datetime
from typing import Any, TYPE_CHECKING, TypedDict
import openpyxl
from config import (CATEGORY_DB_FILE, CATEGORY_DB_TYPE,
                    PRODUCTS_DB_FILE, PRODUCTS_DB_TYPE,
                    ORDER_DB_FILE, ORDER_DB_TYPE)
from provider import provide_db


if TYPE_CHECKING:
    from data_access import DataTyping


Metrics = TypedDict('Metrics', {
    'Total revenue': float,
    'Total amount of products sold': int,
    'Most popular category': str,
    'Most popular product': str
})
category_db = provide_db(CATEGORY_DB_FILE, CATEGORY_DB_TYPE)
products_db = provide_db(PRODUCTS_DB_FILE, PRODUCTS_DB_TYPE)
orders_db = provide_db(ORDER_DB_FILE, ORDER_DB_TYPE)


class Category:
    """A class for working with categories.

    Contains methods for creating an object of class "Category",
    search and write information to the database, work with
    information from the database.
    """
    def __init__(self, name: str) -> None:
        self.cat_name = name
        self.parameters: list[str] = []

    def get_parameters(self) -> 'DataTyping':
        """Returns the list of category parameters from the database.

        :rtype: list[dict[str, str | list[str]]
        :return: the list of category parameters from the database
        """
        data = category_db.read_data()
        result: 'DataTyping' = []
        if data != []:
            for row in data:
                if row.get('category') == self.cat_name.lower():
                    result.append(row)
        return result

    def add_parameters(self) -> None:
        """Adds entered parameters into database."""
        recorded_value: list[dict[str, str | list[str]]] = [{
            'category': self.cat_name,
            'parameters': self.parameters
        }]
        category_db.write_to_data(recorded_value, mode='a')

    def update_category(self, parameters: list[str]) -> None:
        """Updates parameters of existing category.

        :param parameters: list of parameters of category
        :type parameters: list[str]"""
        data: DataTyping = category_db.read_data()
        new_data = {'category': self.cat_name, 'parameters': parameters}
        result_data = []
        for row in data:
            if row.get('category') == self.cat_name:
                result_data.append(new_data)
            else:
                result_data.append(row)
        category_db.write_to_data(result_data)

    def __call__(
            self,
            ) -> None:
        self.add_parameters()


class Product(Category):
    """A class for working with products.

    Contains methods for creating an object of class "Product",
    search and write information to the database, work with
    information from the database.
    """
    def __init__(
            self, category_name: str = ''
            ) -> None:
        super().__init__(name=category_name)
        self.param_values: dict = dict()
        self.id: int = self.generate_id()
        self.quantity: int = 0

    def generate_id(self) -> int:
        """Generates product id based on the information in the database.

        :rtype: int
        :return: unique product id (ordinal number).
        """
        data: DataTyping = products_db.read_data()
        cur_id: int = len(data)
        return cur_id

    def add_product_to_db(self) -> None:
        """Serializes and writes data about product to the database."""
        formatted_data = {'id': self.id,
                          'category': self.cat_name}
        formatted_data = formatted_data | self.param_values
        formatted_data['quantity'] = self.quantity
        formatted_data['created_at'] = time()
        formatted_data['updated_at'] = time()
        exist_data: DataTyping = products_db.read_data()
        new_data: list = []
        for row in exist_data:
            flag = False
            if row.get('category') == self.cat_name:
                flag = True
                for keys, values in self.param_values.items():
                    if row.get(keys) != values:
                        flag = False
            if flag:
                formatted_data['created_at'] = row.get('created_at')
                formatted_data['quantity'] = ((row.get('quantity', 0))
                                              + self.quantity)
                self.id = row.get('id', 0)
                formatted_data['id'] = self.id
            else:
                new_data.append(row)
        new_data.append(formatted_data)
        products_db.write_to_data(new_data)

    def get_all_products(
            self,
            start_date: float,
            end_date: float
            ) -> list[str]:
        """Returns information about all products in database.

        :param start_date: start date of searched period
        :type start_date: float
        :param end_date: end date of searched period
        :type end_date: float

        :rtype: list[dict[str, str]]
        :return: information about all products in database
        """
        exist_data: DataTyping = products_db.read_data()
        result_data: list = []
        for row in exist_data:
            created_time = row.get('created_at', 0.0)
            if start_date <= float(created_time) <= end_date:
                added_data = ''
                if self.cat_name != '':
                    if row.get('category') == self.cat_name.lower():
                        added_data = f"'id': {row.get('id')}, 'name': " \
                                     f" {row.get('name')} \n"
                else:
                    added_data = f"'id': {row.get('id')}, " \
                                 f"'name': {row.get('name')}, " \
                                 f"'category': {row.get('category')} \n"
                if added_data == '':
                    continue
                else:
                    result_data.append(added_data)
        if result_data == []:
            result_data.append('There are no products of the specified '
                               'category added during the specified period '
                               'in the database.')
        return result_data

    def get_product_info(
            self,
            product_id: int,
            quant: int = 0
            ) -> dict[str, str | int | float | bool]:
        """Generates and returns information about concrete product.

        :param product_id: id of searched product
        :type param: int
        :param quant: quantity of product. By default = 0. Needed to check if
                      the product is available for purchase.
        :type quant: int

        :rtype: list[dict[str, str | int | float], str]
        :return: info. about the product and if it is available for purchase.
        """
        data: DataTyping = products_db.read_data()
        new_data: dict[str, str | str | int | float | bool] = dict()
        for row in data:
            if row.get('id') == product_id:
                new_data = row
                new_data.pop('created_at')
                new_data.pop('updated_at')
                new_data['available'] = self.available_product(new_data,
                                                               quant)
        return new_data

    def available_product(self, data: dict[str, str | str | int | float],
                          quant: int = 0) -> bool:
        available_quant = float(data.get('quantity', 0))
        if available_quant < quant:
            return False
        else:
            return True

    def update_quantity(
            self,
            product_id: int,
            new_quantity: int) -> None:
        """Updates quantity of product in the database after purchase.

        :param product_id: id of updated product
        :type param: int
        :param new_quantity: quantity of products purchased
        :type new_quantity: int
        """
        exist_data: DataTyping = products_db.read_data()
        new_data: list = []
        for row in exist_data:
            if row['id'] == product_id:
                row['quantity'] = row.get('quantity', 0) - new_quantity
                row['updated_at'] = time()
            new_data.append(row)
        products_db.write_to_data(new_data)

    def get_category_of_product(self, prod_id: int) -> str:
        """Returns category name of searched product.

        This function is used only for existing products id already known.
        :param prod_id: id of searched product
        :type prod_id: int

        :rtype: str
        :return: category name of product
        """
        data: DataTyping = products_db.read_data()
        result = ''
        for row in data:
            if row.get('id') == prod_id:
                result = row.get('category', '')
        return result


class Order(Product):
    """A class for working with products.

    Contains methods for creating an object of class "Product",
    search and write information to the database, work with
    information from the database.
    """
    def __init__(self, category_name: str = '') -> None:
        super().__init__(category_name)
        self.order_dict: dict[int, int] = {}

    def check_available_product(
            self,
            order_dict: dict[int, int]
    ) -> list[str]:
        res_list = []
        result: str = ('-------------------------------------------\n'
                       '             *** Order info ***\n'
                       '-------------------------------------------\n')
        self.order_dict = order_dict
        for keys, value in self.order_dict.items():
            data = self.get_product_info(product_id=keys, quant=value)
            if data == dict():
                res_list.append('NO')
                result += (f'There are no product with ID "{keys}"'
                           f'in the database.')
                res_list.append(result)
                return res_list
            elif not data['available']:
                res_list.append('NO')
                available_quantity = int(data.get('quantity', 0))
                result += (f'The product with ID {keys} is not '
                           f'available for purchase. \n'
                           f'(Required quantity - {value}, available '
                           f'quantity - {available_quantity}.)')
                res_list.append(result)
                return res_list
            else:
                product_id = keys
                product_name: str = str(data.get('name', ''))
                price: float = float(data.get('price', 0.0))
                quantity = value
                total_price = price * quantity
                result += (f' Product_ID: {product_id} \n'
                           f' Name: {product_name} \n'
                           f' Price: {price} \n'
                           f' Quantity: {quantity} \n'
                           f' Total : {total_price}\n'
                           f'-------------------------------------------\n')
        res_list.append('YES')
        res_list.append(result)
        return res_list

    def gen_order_info(
            self,
            order_dict: dict[int, int]
    ) -> list[dict[str, int | str | float]]:
        self.order_dict = order_dict
        result: list[dict[str, int | str | float]] = []
        for keys, value in self.order_dict.items():
            data = self.get_product_info(product_id=keys, quant=value)
            if data != dict() and data['available']:
                product_id = keys
                product_name = str(data.get('name', ''))
                price = float(data.get('price', 0.0))
                quantity = value
                total_price = price * quantity
                product_order: dict[str, int | str | float] = {
                    'product_id': product_id,
                    'product_name': product_name,
                    'product_price': price,
                    'buy_quantity': quantity,
                    'total_for_product': total_price
                    }
                result.append(product_order)
        return result

    def create_order(
            self,
            product_order: list[dict[str, Any]]
            ) -> None:
        """Serializes and writes data about order to the database.

        :param product_order: data about products in order
        :type product_order: list[dict[str, Any]]
        """
        exist_orders_data: DataTyping = orders_db.read_data()
        order_id = len(exist_orders_data)
        total_amount: float = 0.0
        for row in product_order:
            product_price: float = row.get('total_for_product', 0.0)
            total_amount += product_price
            pr_id: int = row.get('product_id', 0.0)
            buy_quant: int = row.get('buy_quantity', 0)
            self.update_quantity(pr_id, buy_quant)
        formatted_data = [{
            'id': order_id,
            'products_in_order': product_order,
            'total_amount': total_amount,
            'created_at': time()
            }]
        orders_db.write_to_data(formatted_data, mode='a')


class Statistic(Order):
    """A class for creating statistics.

    Contains methods for creating an object of class "Statistics",
    search information in the databases, serializing information and
    writing them to an excel file.
    """
    def __init__(
            self,
            min_date: float,
            max_date: float,
            category_name: str = '',
            ) -> None:
        self.min_date = min_date
        self.max_date = max_date
        super().__init__(category_name)
        self.wb = openpyxl.Workbook(write_only=False)
        self.sheet_category = self.wb.active
        self.sheet_products = self.wb.create_sheet('Products')
        self.sheet_orders = self.wb.create_sheet('Orders')
        self.sheet_metrics = self.wb.create_sheet('Metrics')
        self.orders_stat: dict[int, float] = {int(
            row.get('id', 0)):
            float(row.get('total_amount', 0.0))
            for row in orders_db.read_data()}
        self.product_stat: dict[str, list[int]] = {str(
            row.get('name', '')): [0, 0] for row in products_db.read_data()}
        self.categories_stat: dict[str, list[int]] = {str(
            row.get('category', '')):
            [0, 0] for row in category_db.read_data()}
        self.metric: Metrics = {
            'Total revenue': 0,
            'Total amount of products sold': 0,
            'Most popular category': '',
            'Most popular product': ''
        }
        self.file_name = ''

    def create_stat(self) -> None:
        """ Generates statistics data.

        Updates the class parameters necessary for generating statistics.
        """
        orders_data = orders_db.read_data()
        for order in orders_data:
            date_creating = float(order.get('created_at', 0.0))
            if self.min_date <= date_creating <= self.max_date:
                self.metric['Total revenue'] += order.get('total_amount', 0)
                order_id = int(order.get('id', 0))
                self.orders_stat[order_id] = float(
                    order.get('total_amount', 0.0))
                prod_in_order = list(order.get('products_in_order', []))
                for product in prod_in_order:
                    self.metric['Total amount of '
                                'products sold'] += product.get('buy_quantity')
                    self.product_stat[product.get(
                        'product_name')][0] += product.get('buy_quantity', 0)
                    self.product_stat[product.get(
                        'product_name')][1] += product.get(
                        'total_for_product', 0)
                    category_name = self.get_category_of_product(
                        prod_id=product.get('product_id'))
                    self.categories_stat[category_name][0] += product.get(
                        'buy_quantity')
                    self.categories_stat[category_name][1] += product.get(
                        'total_for_product')
        sorted_orders_stat = sorted(
            self.orders_stat.items(), key=lambda item: item[1], reverse=True)
        self.orders_stat = dict(sorted_orders_stat)
        sorted_prod_stat = sorted(
            self.product_stat.items(), key=lambda item: item[1][0],
            reverse=True)
        self.metric['Most popular product'] = sorted_prod_stat[0][0]
        sorted_categories_stat = sorted(
            self.categories_stat.items(), key=lambda item: item[1][0],
            reverse=True)
        self.metric['Most popular category'] = sorted_categories_stat[0][0]

    def create_category_sheet(self) -> None:
        """Serializes and writes data to an Excel sheet "Category". """
        self.sheet_category.title = 'Categories'
        self.sheet_category.append(['Category name',
                                    'Units sold',
                                    'Total revenue'])
        for row in self.categories_stat.items():
            self.sheet_category.append([row[0], row[1][0], row[1][1]])
        self.sheet_category.column_dimensions['A'].width = 15
        self.sheet_category.column_dimensions['B'].width = 10
        self.sheet_category.column_dimensions['C'].width = 15

    def create_products_sheet(self) -> None:
        """Serializes and writes data to an Excel sheet "Products". """
        self.sheet_products.append(
            ['Product name',
             'Units sold',
             'Total revenue']
            )
        for row in self.product_stat.items():
            self.sheet_products.append([row[0], row[1][0], row[1][1]])
        self.sheet_products.column_dimensions['A'].width = 15  # type: ignore
        self.sheet_products.column_dimensions['B'].width = 10  # type: ignore
        self.sheet_products.column_dimensions['C'].width = 15  # type: ignore

    def create_orders_sheet(self) -> None:
        """Serializes and writes data to an Excel sheet "Orders". """
        self.sheet_orders.append(
            ['Order ID',
             'Total revenue']
        )
        for row in self.orders_stat.items():
            self.sheet_orders.append([row[0], row[1]])
        self.sheet_orders.column_dimensions['A'].width = 15  # type: ignore
        self.sheet_orders.column_dimensions['B'].width = 15  # type: ignore

    def create_metrics_sheet(self) -> None:
        """Serializes and writes data to an Excel sheet "Metrics". """
        for row in self.metric.items():
            self.sheet_metrics.append(
                [row[0], row[1]]
                )
        self.sheet_metrics.column_dimensions['A'].width = 30  # type: ignore
        self.sheet_metrics.column_dimensions['B'].width = 15  # type: ignore

    def save_file(self) -> None:
        """Saves the Excel file with previously recorded data."""
        start_date = 'start: '
        end_date = 'end: '
        if self.min_date == 0:
            start_date += 'no_limit'
        else:
            date_start = datetime.fromtimestamp(self.min_date)
            start_date += (f'{str(date_start.year)}-'
                           f'{str(date_start.month)}-'
                           f'{str(date_start.day)} 00:00:00')
        if self.max_date >= time() - 5:
            end_date += 'no_limit'
        else:
            date_end = datetime.fromtimestamp(self.max_date)
            end_date += (f'{str(date_end.year)}-'
                         f'{str(date_end.month)}-'
                         f'{str(date_end.day)} 00:00:00')
        self.file_name = (f'statistics({start_date}; '
                          f'{end_date}).xlsx')
        self.wb.save(f'./data_access/{self.file_name}')

    def __call__(self) -> None:
        """When the class object is called, statistics are generated,
        data are generated and written into the sheets of the Excel file,
        and the file is saved."""
        self.create_stat()
        self.create_category_sheet()
        self.create_products_sheet()
        self.create_orders_sheet()
        self.create_metrics_sheet()
        self.save_file()
        print(f'The statistics are generated. Check '
              f'the "{self.file_name}" file in the data_access folder.')
