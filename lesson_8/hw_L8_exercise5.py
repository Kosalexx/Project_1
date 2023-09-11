import csv
import openpyxl
import string

wb = openpyxl.Workbook()
wb.create_sheet('Page1', 0)
wb.save('data_file_ex5.xlsx')
wb = openpyxl.load_workbook('data_file_ex5.xlsx')
wb.create_sheet('Page1')
worksheet = wb['Page1']

with open('data_file_ex4.csv', encoding='utf-8') as csv_file:
    file_reader = csv.reader(csv_file)
    my_data = []
    for row in file_reader:
        my_data.append(row)

letters = string.ascii_uppercase
for i in range(len(my_data)):
    for c in range(len(my_data[i])):
        if c < 2:
            worksheet[f'{letters[i]}{c + 2}'] = my_data[i][c]  # type: ignore
        elif c > 2:
            worksheet[f'{letters[i]}{c + 1}'] = my_data[i][c]  # type: ignore

wb.save('data_file_ex5.xlsx')
